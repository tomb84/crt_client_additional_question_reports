# -*- coding: utf-8 -*-
"""
Created on Mon Sep 26 16:12:38 2022

@author: TomBreider
"""



import pandas
import numpy as np
from openpyxl import load_workbook
import os.path
import time
import re
import sys
from openpyxl import load_workbook
import datetime as dt
from pandas.tseries.offsets import DateOffset  

#Option1. 
import glob
import os
from timeit import default_timer as timer

def operation_time(start,operation):
    end = timer()
    print('time for operation: ',operation,end - start)
    start=timer()
    
    return start


def get_question_label_dictionary():

    #Read in the question dictionary
    print("Opening cRT AddQ dictionary") 
    df_dict = pandas.read_excel('CRT AddQ dictionary test.xlsx') #, encoding="utf-8")
    
    #Keep only client additional question rows
    df_dict2 = df_dict[df_dict['Variable'].str.contains('Q_N|Q_L')]
    #Clean the variable name from the label
    df_dict2['Label2'] = df_dict2['Label'].apply(lambda x : ' '.join(x.split(' ')[1:]))
    df_dict2=df_dict2[['Variable','Label2']]
    df_dict2.rename(columns = {'Label2':'Question Label'}, inplace = True)
    df_dict2.rename(columns = {'Variable':'Question'}, inplace = True)
    print("Completed cleading of cRT AddQ dictionary") 
    
    return df_dict2

def scores_to_ovelay(Overlay_Scores):
    
    if(Overlay_Scores == 'Reputation'):
        score_cols = ['Reputation_Score']
    
    if(Overlay_Scores == 'Drivers'):
       score_cols = ['Products','Innovation','Workplace','Citizenship','Conduct','Leadership','Performance']
    
    if(Overlay_Scores == 'Factors'):
       score_cols = ['high_quality_products_and_services',
     'good_value_products_and_services',
     'Stands_behind_products_and_services',
     'Meets_customer_needs',
     'Is_an_innovative_company',
     'Is_generally_first_company_to_go_to_market',
     'Adapts_quickly_to_change',
     'Offers_equal_opportunities_in_the_workplace',
     'Rewards_employees_fairly',
     'Demonstrates_concern_for_the_health_and_wellbeing_of_employees',
     'fair_in_the_way_it_does_business',
     'Behaves_ethically',
     'Open_and_transparent_about_the_way_the_company_operates',
     'Acts_responsibly_to_protect_the_environment',
     'Has_a _positive_influence_on_society',
     'Supports_good_causes',
     'Well_organized_company',
     'Strong_and_appealing_leader',
     'Excellent_managers',
     'Clear_vision_for_its_future',
     'Profitable_company',
     'Strong_prospects_for_future_growth',
     'Delivers_financial_results_that_are_better_than_expected']
    
    return score_cols

def read_data_file(start,input_file):
    
    print("Opening _input_file: ",input_file)
    data = pandas.read_csv(input_file, encoding="utf-8")
    operation_time(start,'open file')
    
    return start,data

def clean_score_col_names(dfx):
    #Clean the score vars
    dfx.columns = dfx.columns.str.replace("_G_reb", "")
    dfx.columns = dfx.columns.str.replace("_G_Reb", "")
    dfx.rename(columns = {'Pulse':'Reputation_Score'}, inplace = True)    
    dfx.rename(columns = {'Governance':'Conduct'}, inplace = True)
    dfx.rename(columns = {'Q320_1r': 'high_quality_products_and_services'}, inplace = True)
    dfx.rename(columns = {'Q320_2r': 'good_value_products_and_services'}, inplace = True)
    dfx.rename(columns = {'Q320_3r': 'Stands_behind_products_and_services'}, inplace = True)
    dfx.rename(columns = {'Q320_4r': 'Meets_customer_needs'}, inplace = True)
    dfx.rename(columns = {'Q320_5r': 'Is_an_innovative_company'}, inplace = True)
    dfx.rename(columns = {'Q320_6r': 'Is_generally_first_company_to_go_to_market'}, inplace = True)
    dfx.rename(columns = {'Q320_7r': 'Adapts_quickly_to_change'}, inplace = True)
    dfx.rename(columns = {'Q320_8r': 'Offers_equal_opportunities_in_the_workplace'}, inplace = True)
    dfx.rename(columns = {'Q320_9r': 'Rewards_employees_fairly'}, inplace = True)
    dfx.rename(columns = {'Q320_10r': 'Demonstrates_concern_for_the_health_and_wellbeing_of_employees'}, inplace = True)
    dfx.rename(columns = {'Q320_11r': 'fair_in_the_way_it_does_business'}, inplace = True)
    dfx.rename(columns = {'Q320_12r': 'Behaves_ethically'}, inplace = True)
    dfx.rename(columns = {'Q320_13r': 'Open_and_transparent_about_the_way_the_company_operates'}, inplace = True)
    dfx.rename(columns = {'Q320_14r': 'Acts_responsibly_to_protect_the_environment'}, inplace = True)
    dfx.rename(columns = {'Q320_15r': 'Has_a _positive_influence_on_society'}, inplace = True)
    dfx.rename(columns = {'Q320_16r': 'Supports_good_causes'}, inplace = True)
    dfx.rename(columns = {'Q320_17r': 'Well_organized_company'}, inplace = True)
    dfx.rename(columns = {'Q320_18r': 'Strong_and_appealing_leader'}, inplace = True)
    dfx.rename(columns = {'Q320_19r': 'Excellent_managers'}, inplace = True)
    dfx.rename(columns = {'Q320_20r': 'Clear_vision_for_its_future'}, inplace = True)
    dfx.rename(columns = {'Q320_21r': 'Profitable_company'}, inplace = True)
    dfx.rename(columns = {'Q320_22r': 'Strong_prospects_for_future_growth'}, inplace = True)
    dfx.rename(columns = {'Q320_23r': 'Delivers_financial_results_that_are_better_than_expected'}, inplace = True)
    
    return dfx

def clean_up_company_column(dfx):
    dfx['Company'] = dfx['Company'].replace(to_replace= r'\\', value= '', regex=True)
    dfx['Company'] = dfx['Company'].replace(to_replace= r'\/', value= '', regex=True)
    dfx['Company'] = dfx['Company'].str.encode("ascii", "ignore")
    
    return dfx

def drop_empty_columns(start,df_tmpx):
    #drop empty columns
    nan_value = float("NaN")
    df_tmpx.replace(" ", nan_value, inplace=True)
    df_tmpx.dropna(how='all', axis=1, inplace=True)
    start=operation_time(start,'replace nans')  
    
    return start,df_tmpx

def drop_inactive_client_additional_cols(df_tmpx):        
    #Now drop the unactive columns
    df_tmpx['Date2'] = pandas.to_datetime(df_tmpx['Month'])
    most_recent_month = df_tmpx['Date2'].max()
    #Get data only for the most recent month
    df_active = df_tmpx[df_tmpx['Date2']==most_recent_month]
    #Now drop the empty cols
    empty_cols = [col for col in df_active.columns if df_active[col].isnull().all()]
    df_tmpx.drop(empty_cols, axis=1,inplace=True)
    
    return df_tmpx
    
def add_score_col_for_each_score(df_tmp2x,score_cols):
    #Now add a _w col for each score col
    for score in score_cols:
        df_tmp2x[score] = df_tmp2x[score].astype(float)
        df_tmp2x[score+'_w']=df_tmp2x[score]*df_tmp2x['Weight_Final']
           
    #score_cols_w = [x + '_w' for x in score_cols] 
    
    #Drop orig score cols
    for score in score_cols:
        df_tmp2x.drop(columns = {score}, inplace = True)
            
    for score in score_cols:
        df_tmp2x.rename(columns = {score+'_w':score}, inplace = True)  
        
    return df_tmp2x   

def add_cols_for_data_cross_cuts(c,Q_N_colsx,df_tmp2x):
                
    ##Warning ----- Client specific code 
    #Vodaphone is out of control - It looks like empty questions in different markets might cause bugs
    if(c == 'Vodafone'): Q_N_colsx = ['Q_N_1283_005_1','Q_N_1283_006_1']
    
    # Basically copy the Q_addit_cols
    for Q_N in Q_N_colsx:
        df_tmp2x[Q_N +'_cross_cut']=df_tmp2x[Q_N]
        
    return Q_N_cols, df_tmp2x 
    
def update_demographic_labels_for_cross_cut_variables(df_demo_mergedx,df_dict2):   
    #Now clean the cross cut cells
    df_demo_mergedx['Demographic'] = df_demo_mergedx['Demographic'].str.replace(r'_cross_cut', '')
    
    #Now merge in the question label into the demographic column
    df_dict3=df_dict2.copy()
    df_dict3.rename(columns = {'Question':'Demographic'}, inplace = True)
    df_dict3['Question Label Merged']=df_dict3['Demographic']+':'+df_dict3['Question Label']
    df_dict3.drop(columns = 'Question Label', inplace = True)
    
    df_demo_mergedx=df_demo_mergedx.merge(df_dict3,how='left', on=['Demographic'])   
    df_demo_mergedx['Demographic_store']=df_demo_mergedx['Demographic']
    df_demo_mergedx.loc[df_demo_mergedx["Demographic"].isin(list(Q_N_cols)),'Demographic'] = df_demo_mergedx['Question Label Merged']
    
    return df_demo_mergedx


#This code solves the problem where all months are are different shapes
def create_a_common_dataframe_for_all_months_data(start,df_add_q_compilex,df_add_q_compx):
    
    #Now we need to create a df with all the empty rows in
    df_add_q_master = df_add_q_compx.groupby(['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Question Label','Answer','Overlay Variable'], as_index=False)['Count'].sum()
    df_add_q_master.drop(columns = {'Count'}, inplace = True)
    
    #Now loop over the months and left merge the results into the master df
    months_in_data = list(df_add_q_compx['Month'].unique())
    for mon in months_in_data:
        
        #Select time period
        df_add_q_comp_atmp = df_add_q_compx.copy()
        df_add_q_comp_atmp2=df_add_q_comp_atmp[df_add_q_comp_atmp['Month']==mon]
        df_add_q_comp_atmp2.drop(columns = {'Question Label'}, inplace = True)
        
        #Must be a left merge - This wil only add results for questions that were asked in the most recent month.
        df_add_q_master_tmp=df_add_q_master.merge(df_add_q_comp_atmp2, on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'], how='left')

        #Now fix the percent col
        df_add_q_master_tmp['Month']=mon               
        df_add_q_compilex = df_add_q_compilex.append(df_add_q_master_tmp)
                                                      
    start=operation_time(start,'month consistency merge') 
    
    return start,df_add_q_compilex

def set_empty_percent_to_zero(df_add_q_compilex):
    
    df_add_q_compilex['Question Count']=df_add_q_compilex.groupby(['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Question'])['Count'].transform('sum')
    #fix_empty=[(df_add_q_compilex["Percent"].isna()) & (df_add_q_compilex["Question Count"].isna())]
    df_add_q_compilex.loc[(df_add_q_compilex["Percent"].isna()) & (df_add_q_compilex["Question Count"]>=0.1),'Percent'] = 0.0
    
    return df_add_q_compilex

def add_quarterly_scores(start,df_add_q_compilex):
    
    #Now groupby quarter
    grp_by_quarter_id = ['STAKEHOLDER_UNIQUE','COUNTRY','Company','Quarter','Demographic','Profile','Question','Answer','Overlay Variable']
    df_demo_p = df_add_q_compilex.groupby(grp_by_quarter_id, as_index=False)['Percent'].mean()
    df_demo_s = df_add_q_compilex.groupby(['STAKEHOLDER_UNIQUE','COUNTRY','Company','Quarter','Demographic','Profile','Question','Answer','Overlay Variable'], as_index=False)['Score'].mean()
    df_demo_c = df_add_q_compilex.groupby(['STAKEHOLDER_UNIQUE','COUNTRY','Company','Quarter','Demographic','Profile','Question','Answer','Overlay Variable'], as_index=False)['Count'].sum()
    df_demo_quarter=df_demo_p.copy()
    df_demo_quarter=df_demo_quarter.merge(df_demo_s,how='left', on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Quarter','Demographic','Profile','Question','Answer','Overlay Variable'])
    df_demo_quarter=df_demo_quarter.merge(df_demo_c,how='left', on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Quarter','Demographic','Profile','Question','Answer','Overlay Variable'])

    df_demo_quarter.rename(columns = {'Quarter':'Month'}, inplace = True)

    start=operation_time(start,'add quarterly scores') 
    
    return start,df_demo_quarter

def add_rolling_scores(start,df_add_q_compilex,nmonths_data):
            
    #Now we can add the rolling scores
    df_roll = df_add_q_compilex.copy()  
    grp_by_roll_id = ['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable']

    #add a check to make sure that max month = real current month -1 
    df_roll_compile=pandas.DataFrame()
    
    most_recent_month = df_roll['Date2'].max()
    most_recent_month_str= df_roll.loc[df_roll.Date2==most_recent_month, 'Month'].values[0]
    
    #Now build the rolling scores for 3,6,9,12 months
    for mroll in range(3,13,3):
        #print(mroll)
        
        rollmin_month = most_recent_month + DateOffset(months=-(mroll-1))
        #print('rollmin_month=',rollmin_month)
        
        #if we have data for greater than rolling period months
        if(nmonths_data >= mroll):
            df_roll_tmp = df_roll.copy()
            df_roll_tmp=df_roll_tmp[df_roll_tmp['Date2']>=rollmin_month]
            #print(len(df_roll_tmp['Date2'].unique()))
            df_roll_tmp_p = df_roll_tmp.groupby(grp_by_roll_id, as_index=False)['Percent'].mean()
            df_roll_tmp_s = df_roll_tmp.groupby(grp_by_roll_id, as_index=False)['Score'].mean()
            df_roll_tmp_c = df_roll_tmp.groupby(grp_by_roll_id, as_index=False)['Count'].sum()
            df_roll_results = df_roll_tmp_p.copy()
            df_roll_results=df_roll_results.merge(df_roll_tmp_s,how='left', on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'])
            df_roll_results=df_roll_results.merge(df_roll_tmp_c,how='left', on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'])
            
            df_roll_results['Month']=str(mroll)+'month rolling'
    
            df_roll_compile=df_roll_compile.append(df_roll_results)
            
    start=operation_time(start,'add rolling scores') 
    
    return start,df_roll_compile,most_recent_month,most_recent_month_str

def reshape_dataframe_put_month_quarter_and_rollups_into_columns(start,df_add_q_compilex,df_add_q_compile_reshapex):
    
            time_periods = list(df_add_q_compilex['Time Period'].unique())
            for tp in time_periods:
               # print('Reshaping output frame:',tp)
                
                #Select time period
                df_add_q_compile_tmp = df_add_q_compilex.copy()
                df_add_q_compile_tmp2=df_add_q_compile_tmp[df_add_q_compile_tmp['Time Period']==tp]
                
                #Now add time period into col names
                df_add_q_compile_tmp2.rename(columns = {'Percent':str(tp)+' %'}, inplace = True)
                df_add_q_compile_tmp2.rename(columns = {'Count':str(tp)+' Count'}, inplace = True)
                df_add_q_compile_tmp2.rename(columns = {'Score':str(tp)+' Score'}, inplace = True)
                
                #now drop time period from
                df_add_q_compile_tmp2.drop(columns = {'Time Period'}, inplace = True)
                df_add_q_compile_tmp2.drop(columns = {'Question Label'}, inplace = True)
                
                #Must be a left merge - This wil only add results for questions that were asked in the most recent month.
                df_add_q_compile_reshapex=df_add_q_compile_reshapex.merge(df_add_q_compile_tmp2, on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'], how='left')
                
                #print('Appening time_period: ',tp,df_add_q_compile_reshape.shape)

            start=operation_time(start,'reshape of rows into cols') 
            
            return start, df_add_q_compile_reshapex
        
def export_results_to_excel(start,cx,most_recent_month_strx,Overlay_Scoresx,df_add_q_compile_reshapex):
                
    excel_sheet = 'Results'
    
    time.sleep(1.0)     
    #Test if global file exits
    output_file='CRT Client Additional Questions output '+cx+' '+most_recent_month_strx+'.xlsx'
    if(Overlay_Scoresx != 'Reputation'):
        output_file='CRT Client Additional Questions output '+cx+' '+most_recent_month_strx+' '+Overlay_Scoresx+' score overlays.xlsx'  
                        
    if os.path.isfile(output_file):
        #print ("File exists - append sheet to global scores file")
        workbook = load_workbook(output_file)
        writer = pandas.ExcelWriter(output_file, engine = 'openpyxl')    
        writer.book = workbook
        df_add_q_compile_reshapex.to_excel(writer, sheet_name=excel_sheet,index=False) #,float_format="%.2f"
        worksheet = writer.sheets[excel_sheet]
        #Set columns formats - BUG - Does not work from some reason    
        writer.save() 
        writer.close()      
    else:
        #print ("File does not exist - create company scores file")
        writer = pandas.ExcelWriter(output_file, engine = 'xlsxwriter')
        df_add_q_compile_reshapex.to_excel(writer,sheet_name=excel_sheet,index=False)
        workbook  = writer.book
        worksheet = writer.sheets[excel_sheet]
        #Set columns formats
        format_percent = workbook.add_format({'num_format': '0.0%'})
        format_score = workbook.add_format({'num_format': '0.0'})
        format_count = workbook.add_format({'num_format': '0'})
        for pindex in range(9, df_add_q_compile_reshapex.shape[1], 3):    
            worksheet.set_column(pindex, pindex, 4, format_percent)
        for cindex in range(10, df_add_q_compile_reshapex.shape[1], 3):
            worksheet.set_column(cindex, cindex, 4, format_count)
        for sindex in range(11, df_add_q_compile_reshapex.shape[1], 3):
            worksheet.set_column(sindex, sindex, 4, format_score) 
        writer.save() 
        writer.close() 
        
        start=operation_time(start,'Write output complete')
        print("Output complete for company: ",c)   
        
        return start

def get_2022_country_codebook_dictionary():

    #Read in the question dictionary
    print("Opening codebook") 
    df_codebookx = pandas.read_csv('Country Codebook - 2022 - 16Feb22.csv')
    
    df_codebookx['Year']=2022
    df_codebookx.rename(columns = {'Country':'COUNTRY'}, inplace = True)
    df_codebookx=df_codebookx[['COUNTRY','Year','global_mean','global_stdev','CM_Reb','CSD_Reb']]
    
    return df_codebookx

def get_2021_country_codebook_dictionary():

    #Read in the question dictionary
    print("Opening codebook") 
    df_codebookx = pandas.read_csv('Country Codebook Ages65_02Feb21.csv')
    
    df_codebookx['Year']=2021
    df_codebookx.rename(columns = {'Country':'COUNTRY'}, inplace = True)
    df_codebookx=df_codebookx[['COUNTRY','Year','global_mean','global_stdev','CM_Reb','CSD_Reb']]
    
    return df_codebookx

def clean_likert_scale_vars(start,Q_L_colsx,df_tmpx):
    
    #Clean the likert scale vars- keep only the numbers
    for v in Q_L_colsx:
        print('Cleaning Q_L variable:',v)
        df_tmpx.loc[df_tmpx[v]=='Not sure',v] = '99'
        #df_tmp[v]=df_tmp[v].str.extract('(\d+)')
        #print(df_tmp[v].unique())
        df_tmpx[v]=df_tmpx[v].str.extract('(\d+)')
        #print(df_tmpx[v].unique())
        
        # Alternative approach  - above is faster
        #df_tmp.loc[df_tmp[v]=='Does not describe well 1',v] = 1
        #df_tmp.loc[df_tmp[v]=='1 Strongly disagree',v] = 1
        #df_tmp.loc[df_tmp[v]=='I strongly disagree 1',v] = 1
        #df_tmp.loc[df_tmp[v]=='1 Not at all',v] = 1
        #df_tmp.loc[df_tmp[v]=='1 Not at all important',v] = 1
        #df_tmp.loc[df_tmp[v].isin(list(Q_N_cols)),'Demographic'] = df_demo_mergedx['Question Label Merged']
        
        #CLean the 7s
        #df_tmp.loc[df_tmp[v]=='Describes very well 7',v] = 7
        #df_tmp.loc[df_tmp[v]=='7 Strongly agree',v] = 7
        #df_tmp.loc[df_tmp[v]=='I strongly agree 7',v] = 7
        #df_tmp.loc[df_tmp[v]=='7 Very much',v] = 7
        #df_tmp.loc[df_tmp[v]=='7 Very important',v] = 7
    
        try:
            df_tmpx[v] = df_tmpx[v].astype(float)
            print(v,df_tmp[v].dtypes)  
        except:
            print('Q_L_clean failed:',df_tmpx[v].unique())
            
        start=operation_time(start,'clean likert scale cols')    
            
        return start,df_tmpx
    
def stack_data_file_by_unique_stakeholder(start, dfx):
        
    df_unique_shx = dfx.loc[dfx['Stakeholder_IGP1']=='Yes']
    df_unique_shx['STAKEHOLDER_UNIQUE']='IGP'
    
    sh_colNames = list(df.columns[df.columns.str.contains(pat = 'Stakeholder_')])
    try:
        sh_colNames.remove('Stakeholder_IGP1')
    except:
        print('Colnames do not include Stakeholder_IGP1 - error')
        sys.exit(1)
        
    #Append extra stakeholder data to the end of the df 
    for sh in sh_colNames:
        #First replace any 1 with "Yes"
        df[sh] = df[sh].astype(str)
        df[sh].str.replace('1', 'Yes')
        print(sh,df[sh].value_counts())
        df_sh_tmp = df.loc[df[sh]=='Yes']
        df_sh_tmp['STAKEHOLDER_UNIQUE']=sh
        if(df_sh_tmp.shape[0]>0):
            df_unique_shx=df_unique_shx.append(df_sh_tmp)
    
    #NOW DROP THE ORIGINAL STAKEHOLDER COLS    
    sh_colNames.append('Stakeholder_IGP1')
    df_unique_shx.drop(sh_colNames, axis=1,inplace=True)
    
    df_unique_shx['STAKEHOLDER_UNIQUE']=df_unique_shx['STAKEHOLDER_UNIQUE'].str.replace("Stakeholder_", "")
    df_unique_shx['STAKEHOLDER_UNIQUE']=df_unique_shx['STAKEHOLDER_UNIQUE'].str.replace("1", "")
    
    start=operation_time(start,'stakeholder stacking') 
    
    return start, df_unique_shx  

       
def clean_likert_scale_vars(start, df_tmpx,Q_L_cols):        
    for v in Q_L_cols:
        df_tmpx.loc[df_tmpx[v]=='Not sure',v] = '99'
        df_tmpx[v]=df_tmpx[v].str.extract('(\d+)')
        try:
            df_tmpx[v] = df_tmpx[v].astype(float)
        except:
            print(df_tmpx[v].unique())
            
    start=operation_time(start,'clean likert scale cols')
    
    return start, df_tmpx

def calculate_Q_L_scores(df_tmpx):    
    for v in Q_L_cols:
        df_tmpx[v+'_score']= (((100*((df_tmpx[v]-1)/6)-df_tmpx['CM_Reb'])/df_tmpx['CSD_Reb'])*df_tmpx['global_stdev'])+df_tmpx['global_mean']

    #Now replace high values (99s) with nan
    for v in Q_L_cols:
        nan_value = float("NaN")
        df_tmpx.loc[df_tmpx[v+'_score']>= 300, v+'_score'] = nan_value
        
    return df_tmpx

def recode_likert_scale_vars(start,df_tmpx):            
    for v in Q_L_cols:
        df_tmpx[v] = df_tmpx[v].astype(str)
        df_tmpx[v].dtypes
        df_tmpx.loc[df_tmpx[v]=='1.0', v] = 'B2'
        df_tmpx.loc[df_tmpx[v]=='2.0', v] = 'B2'
        df_tmpx.loc[df_tmpx[v]=='3.0', v] = 'M3'
        df_tmpx.loc[df_tmpx[v]=='4.0', v] = 'M3'
        df_tmpx.loc[df_tmpx[v]=='5.0', v] = 'M3'
        df_tmpx.loc[df_tmpx[v]=='6.0', v] = 'T2'
        df_tmpx.loc[df_tmpx[v]=='7.0', v] = 'T2'
        df_tmpx.loc[df_tmpx[v]=='99.0', v] = 'Not Sure'
        df_tmpx[v].value_counts()
        
    start=operation_time(start,'recode Q_L vars')
    
    return start, df_tmpx

def put_Q_L_score_cols_into_question_col(start, df_tmp4ax,df_tmp4,Q_N_cols, Q_L_cols, Q_L_score_cols):                
    
    df_tmp4ax=df_tmp4ax.loc[df_tmp4ax['Question'].isin(Q_N_cols)]
    
    #Drop the Q_L score cols
    df_tmp4ax.drop(Q_L_score_cols, axis=1,inplace=True)
    
    #Now append the other score vars into the Question score
    for v, value in enumerate(Q_L_cols):
        df_tmp4b=df_tmp4.loc[df_tmp4['Question']==value]
        df_tmp4b=df_tmp4b[['STAKEHOLDER_UNIQUE','COUNTRY','Company',
             'Weight_Final','Month','Demographic','Profile','Question','Answer','Reputation_Score',str(value)+'_score']]
        df_tmp4b.rename(columns = {value+'_score':'Question_Score'}, inplace = True)
        
        df_tmp4ax=df_tmp4ax.append(df_tmp4b)
        
    start=operation_time(start,'put Q_L scores into question col')
    
    return start, df_tmp4ax

def calculate_aggregated_client_question_scores(start, df_demo_mergedx, df_tmp5, grp_by_vars_qst,grp_by_vars_prof, grp_by_vars_demo):
    
    #Keep only reputation data in this table
    df_demo_mergedx=df_demo_mergedx.loc[df_demo_mergedx['Overlay Variable']=='Reputation_Score']            
    
    #Next table re-do above for Question scores but keep only addit question scores data
    df_tmp6=df_tmp5.loc[df_tmp5['Overlay Variable']=='Question_Score']
    
    #Drop the Q_N vars from the Question rows - we do not need scores for Q_N vars
    df_tmp6=df_tmp6.loc[df_tmp6['Question'].isin(Q_L_cols)]
    
    #Calc the aggregated weighted question scores 
    df_demo_qsc = df_tmp6.groupby(grp_by_vars_qst, as_index=False)['Score'].sum()
    df_demo_merged_q = df_demo_qsc.copy()
    df_demo_qscw = df_tmp6.groupby(grp_by_vars_qst, as_index=False)['Weight_Final_valid'].sum()
    df_demo_merged_q=df_demo_merged_q.merge(df_demo_qscw,how='left', on=grp_by_vars_qst)
    df_demo_merged_q['Score'] = df_demo_merged_q['Score']/df_demo_merged_q['Weight_Final_valid']
    df_demo_merged_q['Answer'] = 'Score'
    
    #Now for the counts
    sum_demo_weight = df_tmp6.groupby(grp_by_vars_qst, as_index=False)['Weight_Final'].sum()
    df_demo_merged_q=df_demo_merged_q.merge(sum_demo_weight,how='left', on=grp_by_vars_qst)
    df_demo_merged_q.rename(columns = {'Weight_Final':'Weight_Final_x'}, inplace = True)

    #Now get the percents based on the profile group
    sum_demo_prof_weight=df_tmp6.groupby(grp_by_vars_prof, as_index=False)['Weight_Final'].sum()
    sum_demo_prof_weight.rename(columns = {'Weight_Final':'Weight_Final_y'}, inplace = True)
    df_demo_merged_q=df_demo_merged_q.merge(sum_demo_prof_weight,how='left', on=grp_by_vars_prof)
    
    sum_demo_weight=df_tmp6.groupby(grp_by_vars_demo, as_index=False)['Weight_Final'].sum()
    sum_demo_weight.rename(columns = {'Weight_Final':'Weight_Final_z'}, inplace = True)
    df_demo_merged_q=df_demo_merged_q.merge(sum_demo_weight,how='left', on=grp_by_vars_demo)
    df_demo_merged_q['Percent'] = df_demo_merged_q['Weight_Final_y']/df_demo_merged_q['Weight_Final_z']

    df_demo_mergedx=df_demo_mergedx.append(df_demo_merged_q)
    df_demo_mergedx.reset_index(drop=True)
        
    start=operation_time(start,'Calculate aggregated client question scores') 
    
    return start, df_demo_mergedx

##### MAIN    
#################################### 
#Dir='C:\Current Processing/'
#os.chdir('C:\Current Processing/')


data_filename = 'CRT - Global - Nov21-Oct22 - cumulative working file CLEANED CQ ACTIVE3'
for file in glob.glob(data_filename+' *.csv'):
    print(file)
    

#Reputation / Drivers / Factors
Overlay_Scores = 'Reputation'
demo_cols =['Overall','Gender','cRT_MILLENNIALS','CUSTOMERS','Age','Region','Education','Income']

#Get the question label dict
df_dict2 = get_question_label_dictionary()

#Get the country codebook.
df_codebook_2022= get_2022_country_codebook_dictionary()
df_codebook_2021= get_2021_country_codebook_dictionary()

    
#Get the score cols to overlay
score_cols = scores_to_ovelay(Overlay_Scores)


#Loop over the input files 
for input_file in glob.glob(data_filename+' LABELS 11*.csv'):
    print(input_file)
    
    start = timer()

    start,data=read_data_file(start,input_file)
    
    df=data.copy()
    
    #Now read in the Likter scale values
    input_file_values = input_file.replace('LABELS','VALUES')
    start,data=read_data_file(start,input_file_values)
    df_values=data.copy()
    
    #Now merge 
    df=df.merge(df_values,how='left', on=['Global_ID']) 
    
    start, df_unique_sh = stack_data_file_by_unique_stakeholder(start, df)
    
    ##So the issue is we cannot trust the value labels in the SPSS data file.
    
    df_unique_sh=clean_score_col_names(df_unique_sh)
    
    df_unique_sh=clean_up_company_column(df_unique_sh)
    
    df_unique_sh['Company'] = df_unique_sh['Company'].str.decode(encoding = 'UTF-8') 
  
    company_list = list(df_unique_sh.Company.unique())
    
    #### WARNING Merc Benz ONLY
    #indexMB=company_list.index('AMP')
    #company_list = company_list[indexMB:]
    
    #company_list = ['Banca March']
    #company_list = ['MGM Resorts International']
    
    for c in company_list:
        print("Starting Report for company",c)
        start = timer()
        start_c = timer()
        
        df_company = df_unique_sh[df_unique_sh['Company']==c]
    
        df_tmp = df_company.copy()
    
        start,df_tmp = drop_empty_columns(start,df_tmp)
        
        df_tmp = drop_inactive_client_additional_cols(df_tmp)

        #Find client relevant columns
        Q_N_cols = [col for col in df_tmp.columns if 'Q_N_' in col]
        Q_L_cols = [col for col in df_tmp.columns if 'Q_L_' in col]
        Q_addit_cols = Q_N_cols + Q_L_cols
        
        print('Company = ',c,' ncountries = ',len(df_tmp['COUNTRY'].unique()), 'Q_N_cols = ',len(Q_N_cols) )
        
        #Merge in combined codebook 
        df_codebook_combined = df_codebook_2022.append(df_codebook_2021)
        df_tmp=df_tmp.merge(df_codebook_combined,how='left', on=['COUNTRY','Year']) 
        
        #Clean the likert scale vars
        start, df_tmp = clean_likert_scale_vars(start,df_tmp,Q_L_cols)
        
        if(Overlay_Scores == 'Reputation'):
            df_tmp = calculate_Q_L_scores(df_tmp)

            Q_L_score_cols = [x+'_score' for x in Q_L_cols]
        
            
        #Recode Likert scale vars to B2, M3, etc.
        start, df_tmp = recode_likert_scale_vars(start,df_tmp)
          
        #Only continue if data exists for client
        if(len(Q_addit_cols)==0):
            print("No AddQ data for company",c)
        else:       
            
            df_add_q_comp=pandas.DataFrame()
            
            score_cols_all=score_cols
    
            if(Overlay_Scores == 'Reputation'):
                score_cols_all=score_cols_all+Q_L_score_cols
                
            df_tmp2 = df_tmp[['STAKEHOLDER_UNIQUE','COUNTRY','Company','Weight_Final','Month']+Q_addit_cols+demo_cols+score_cols_all]
        
            for score in score_cols_all:
                df_tmp2[score] = df_tmp2[score].astype(float)
                df_tmp2[score+'_w']=df_tmp2[score]*df_tmp2['Weight_Final']
                    
            #Drop orig score and rename SCORE_w cols to just SCORE 
            df_tmp2.drop(score_cols_all, axis=1,inplace=True)
            for score in score_cols_all:
                df_tmp2.rename(columns = {score+'_w':score}, inplace = True) 
        
            #CROSS-CUTS CODE only add this in for reputation scores
            if(Overlay_Scores == 'Reputation'):
                Q_N_cols,df_tmp2=add_cols_for_data_cross_cuts(c,Q_N_cols,df_tmp2)
  
            #now get the demos into the rows - re-shape demos
            id_vars = ['STAKEHOLDER_UNIQUE',            'COUNTRY',            'Company',
                         'Weight_Final',              'Month']+Q_addit_cols+score_cols_all
            df_tmp3=df_tmp2.melt(id_vars=id_vars)
            df_tmp3.rename(columns = {'variable':'Demographic'}, inplace = True)
            df_tmp3.rename(columns = {'value':'Profile'}, inplace = True)
            
            #Second re-shape addit questions
            id_vars2 = ['STAKEHOLDER_UNIQUE',            'COUNTRY',            'Company',
                         'Weight_Final',              'Month','Demographic','Profile'] +score_cols_all                           
            df_tmp4=df_tmp3.melt(id_vars=id_vars2)
            df_tmp4.rename(columns = {'variable':'Question'}, inplace = True)
            df_tmp4.rename(columns = {'value':'Answer'}, inplace = True)
            
            #drop rows with nans
            nan_value = float("NaN")
            df_tmp4.loc[df_tmp4["Answer"]=='nan','Answer'] = nan_value
            df_tmp4.dropna(subset=['Answer'], inplace=True)
           
            df_tmp4a=df_tmp4.copy()
            
            #Now get the scores from the addit questions into the Question Score column
            if(Overlay_Scores == 'Reputation'):
                start, df_tmp4a = put_Q_L_score_cols_into_question_col(start, df_tmp4a,df_tmp4,Q_N_cols, Q_L_cols, Q_L_score_cols)
                
            #####################################################################
            
            #Third re-shape score cols
            id_vars3 = ['STAKEHOLDER_UNIQUE',            'COUNTRY',            'Company',
                         'Weight_Final',              'Month','Demographic','Profile','Question','Answer']                          
            df_tmp5=df_tmp4a.melt(id_vars=id_vars3)
            df_tmp5.rename(columns = {'variable':'Overlay Variable'}, inplace = True)
            df_tmp5.rename(columns = {'value':'Score'}, inplace = True)

            #Set weight_final_valid to nan is row score does not exist  = this is key for factor scores which have missing values
            nan_value = float("NaN")
            df_tmp5['Weight_Final_valid']=df_tmp5['Weight_Final']
            df_tmp5.loc[df_tmp5["Score"].isna(),'Weight_Final_valid'] = nan_value
                        
            #Now setup the group bys.
            grp_by_vars_ans=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Demographic','Profile','Question','Answer','Overlay Variable']
            grp_by_vars_qst=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Demographic','Profile','Question','Overlay Variable']
            grp_by_vars_demo=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Demographic','Question']
            grp_by_vars_prof=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Demographic','Profile','Question']
               
            df_demo = df_tmp5.groupby(grp_by_vars_ans, as_index=False)['Weight_Final'].sum()
            sum_demo_weight = df_tmp5.groupby(grp_by_vars_qst, as_index=False)['Weight_Final'].sum()
            df_demo_merged=df_demo.copy()
            df_demo_merged=df_demo_merged.merge(sum_demo_weight,how='left', on=grp_by_vars_qst)
        
            #Get the scores
            df_demo_sc = df_tmp5.groupby(grp_by_vars_ans, as_index=False)['Score'].sum()
            #use _valid to solve factor score issue
            df_demo_scw = df_tmp5.groupby(grp_by_vars_ans, as_index=False)['Weight_Final_valid'].sum()
            df_demo_merged=df_demo_merged.merge(df_demo_sc,how='left', on=grp_by_vars_ans)
            df_demo_merged=df_demo_merged.merge(df_demo_scw,how='left', on=grp_by_vars_ans)

            #Now calc the distrib
            df_demo_merged['Percent'] = df_demo_merged['Weight_Final_x']/df_demo_merged['Weight_Final_y']
            
            #update to _valid to solve factor score issue
            df_demo_merged['Score'] = df_demo_merged['Score']/df_demo_merged['Weight_Final_valid']
            
            if(Overlay_Scores == 'Reputation'):
                start, df_demo_merged = calculate_aggregated_client_question_scores(start, df_demo_merged, df_tmp5, grp_by_vars_qst,grp_by_vars_prof, grp_by_vars_demo)

            #Merge in the Question Label
            df_demo_merged=df_demo_merged.merge(df_dict2,how='left', on=['Question'])   
            
            #Clean cross cuts demographic vars
            if(Overlay_Scores == 'Reputation'):
                update_demographic_labels_for_cross_cut_variables(df_demo_merged,df_dict2)
                
            #Clean the columns
            df_demo_merged=df_demo_merged[['STAKEHOLDER_UNIQUE','COUNTRY','Company','Month','Demographic','Profile','Question','Question Label','Answer','Percent','Weight_Final_x','Overlay Variable','Score']]
            
            #rename weight to count
            df_demo_merged.rename(columns = {'Weight_Final_x':'Count'}, inplace = True)
            
            #Append demo data into compiled df
            df_add_q_comp = df_demo_merged.copy()
            
            ###########################     END OF SCORE CALCULATIONS   ################################

            ############################################################################################
            
            df_add_q_compile=pandas.DataFrame()
            
            #Get the dfs into a commmon shape - rows with empty values must be present.
            start,df_add_q_compile=create_a_common_dataframe_for_all_months_data(start,df_add_q_compile,df_add_q_comp)
           
            #Now make sure that the % col contains a 0.0 if the question has data and percent cell is empty   
            df_add_q_compile=set_empty_percent_to_zero(df_add_q_compile)
            
            #Clean the dates and get the QTR
            df_add_q_compile['Date2'] = pandas.to_datetime(df_add_q_compile['Month'])
            df_add_q_compile['Quarter'] = df_add_q_compile['Date2'].dt.to_period('Q')
            
            start,df_demo_quarter=add_quarterly_scores(start,df_add_q_compile)
            
            nmonths_data = len(list(df_add_q_compile['Month'].unique()))
            start,df_roll_compile,most_recent_month,most_recent_month_str=add_rolling_scores(start,df_add_q_compile,nmonths_data)
        
            #This is a bit complicated becuase the stakeholders and countries can be asked in different months which skews the sorting
            #Sort by date after sorting by stakeholder, country, date
            df_add_q_compile=df_add_q_compile.sort_values(by=['STAKEHOLDER_UNIQUE','COUNTRY','Date2'], ascending=False)
            df_add_q_compile=df_add_q_compile.sort_values(by=['Date2'], ascending=False)

            #Drop Quarter and Date
            df_add_q_compile.drop(columns = {'Quarter','Date2'}, inplace = True)       
            
            #This is where we would add the export for Moritz - Moritz wants one file for all clients.
            
            
            #Now append quarterly into the main df
            df_add_q_compile = df_add_q_compile.append(df_demo_quarter)
            
            #Now append rolling into the main df
            if(nmonths_data >= 3):
                df_add_q_compile = df_add_q_compile.append(df_roll_compile)
                
            
            #Now clean the column order
            df_add_q_compile.rename(columns = {'Month':'Time Period'}, inplace = True)
            df_add_q_compile=df_add_q_compile[['STAKEHOLDER_UNIQUE','COUNTRY','Company','Time Period','Overlay Variable','Demographic','Profile','Question','Question Label','Answer','Percent','Count','Score']]
                 
            #Now remove any data pts with counts < 30.
            df_add_q_compile.loc[df_add_q_compile['Count'] < 30, 'Score'] = 'low sample'
            
            #Now find the time period with the most rows
            ndata_in_period = pandas.DataFrame(df_add_q_compile['Time Period'].value_counts())
            #Now re-structure so that vars are in the 
            #Initialize the cumu array with the time period with the highest number of rows.
            df_add_q_compile_reshape=df_add_q_compile[df_add_q_compile['Time Period']==ndata_in_period.idxmax()[0]]
            df_add_q_compile_reshape.drop(columns = {'Time Period','Percent','Question Label', 'Count', 'Score'}, inplace = True)

            #Above question labels come out as Nans so merge in question labels from the most recent month.
            df_active=df_add_q_compile[df_add_q_compile['Time Period']==most_recent_month_str]
            df_active.drop(columns = {'Time Period'}, inplace = True)
            
            #now merge the two toegther
            df_add_q_compile_reshape=df_add_q_compile_reshape.merge(df_active, on=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'], how='left')
            
            df_add_q_compile_reshape=df_add_q_compile_reshape[['STAKEHOLDER_UNIQUE','COUNTRY','Company','Overlay Variable','Demographic','Profile','Question','Question Label','Answer']]

            start=operation_time(start,'add question labels') 
            
            #Now reshape put time data in columns
            start, df_add_q_compile_reshape=reshape_dataframe_put_month_quarter_and_rollups_into_columns(start,df_add_q_compile,df_add_q_compile_reshape)
            
            
            #Now sort the final output
            df_add_q_compile_reshape=df_add_q_compile_reshape.sort_values(by=['STAKEHOLDER_UNIQUE','COUNTRY','Company','Demographic','Profile','Question','Answer','Overlay Variable'], ascending=False)


            #Export results    
            start=export_results_to_excel(start,c,most_recent_month_str,Overlay_Scores,df_add_q_compile_reshape)   
            start_c=operation_time(start_c,'Full report for company:'+c)
        
